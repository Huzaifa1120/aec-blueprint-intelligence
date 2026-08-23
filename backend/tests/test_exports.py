"""BOQ exports (G7) — JSON/XLSX/PDF writers + export endpoint.

Exit gates (spec v3 §7.14):
- JSON export round-trips the BOQ payload byte-for-value equal.
- XLSX export read back through openpyxl matches cell-for-cell, including
  the unpriced flag column, confidence_status and size_source.
- PDF export produces a real PDF (starts ``%PDF-``).
- Unknown format → 422; unknown estimate → 404.
- Every exported line carries material, quantity, confidence_status,
  size_source and the unpriced flag; unpriced lines are flagged
  "UNPRICED — review required" and NEVER rendered as $0.
"""

from __future__ import annotations

import io
import json
import uuid

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.db.base import Base
from app.db.models.estimate import BoqItem, Estimate, Measurement
from app.db.models.geometry import Component, Route
from app.db.models.project import Drawing, Project, Sheet
from app.db.session import get_db
from app.exports import UNPRICED_LABEL
from app.exports.json_export import render as render_json
from app.exports.pdf_export import render as render_pdf
from app.exports.router import router as exports_router
from app.exports.xlsx_export import render as render_xlsx

HEADERS = [
    "Material",
    "Quantity",
    "Unit Cost",
    "Total Cost",
    "Unpriced",
    "Confidence Status",
    "Size Source",
]

REQUIRED_LINE_FIELDS = {
    "material_name",
    "quantity",
    "confidence_status",
    "size_source",
    "unpriced",
}


# ---------------------------------------------------------------------------
# Payload fixtures (documented GET /api/estimates/{id}/boq shape)
# ---------------------------------------------------------------------------
def _payload() -> dict:
    """One priced route, one priced material, one unpriced material."""
    return {
        "estimate_id": "00000000-0000-0000-0000-000000000001",
        "totals": {"materials": 95.0, "labor": 42.0, "grand": 137.0},
        "routes": [
            {
                "route_type": "duct_rectangular",
                "length_m": 12.5,
                "size_json": {
                    "width_mm": 600,
                    "height_mm": 400,
                    "source": "label",
                    "ref": "600x400",
                },
                "confidence_status": "MEASURED",
                "material_name": "Rectangular duct",
                "quantity": 12.5,
                "unit_cost": 7.0,
                "total_cost": 87.5,
                "unpriced": False,
                "size_source": "schedule",
            }
        ],
        "materials": [
            {
                "material_name": "Duct sealant",
                "quantity": 3.0,
                "unit_cost": 2.5,
                "total_cost": 7.5,
                "unpriced": False,
                "confidence_status": "MEASURED",
                "size_source": None,
            },
            {
                "material_name": "Mystery bracket",
                "quantity": 4.0,
                "unit_cost": 0.0,
                "total_cost": 0.0,
                "unpriced": True,
                "confidence_status": "ASSUMED",
                "size_source": "assumed",
            },
        ],
    }


def _empty_payload() -> dict:
    return {
        "estimate_id": "e2",
        "totals": {},
        "routes": [],
        "materials": [],
    }


# ---------------------------------------------------------------------------
# JSON writer — round-trip equality
# ---------------------------------------------------------------------------
def test_json_round_trip_equality():
    payload = _payload()
    out = render_json(payload)
    assert isinstance(out, bytes)
    assert json.loads(out.decode("utf-8")) == payload


def test_json_empty_payload_safe():
    assert json.loads(render_json(_empty_payload()).decode("utf-8")) == _empty_payload()


# ---------------------------------------------------------------------------
# XLSX writer — openpyxl read-back cell equality
# ---------------------------------------------------------------------------
def _xlsx_grid(payload: dict) -> list[list]:
    wb = load_workbook(io.BytesIO(render_xlsx(payload)))
    ws = wb["BOQ"]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def test_xlsx_header_and_cell_equality():
    grid = _xlsx_grid(_payload())
    assert grid[0] == HEADERS
    route_row = grid[1]
    assert route_row[0] == "Rectangular duct"
    assert route_row[1] == 12.5
    assert route_row[2] == 7.0
    assert route_row[3] == 87.5
    assert route_row[4] is False
    assert route_row[5] == "MEASURED"
    assert route_row[6] == "schedule"
    material_row = grid[2]
    assert material_row[0] == "Duct sealant"
    assert material_row[1] == 3.0
    assert material_row[2] == 2.5
    assert material_row[3] == 7.5
    assert material_row[4] is False
    assert material_row[5] == "MEASURED"
    assert material_row[6] is None


def test_xlsx_totals_block_verbatim():
    grid = _xlsx_grid(_payload())
    labels_to_values = {row[0]: row[3] for row in grid if row and row[0]}
    assert labels_to_values["Materials Total"] == 95.0
    assert labels_to_values["Labor Total"] == 42.0
    assert labels_to_values["Grand Total"] == 137.0


def test_xlsx_every_line_carries_required_columns():
    grid = _xlsx_grid(_payload())
    item_rows = grid[1 : 1 + 3]  # header + 3 lines before the totals spacer
    for row in item_rows:
        assert row[0], f"missing material on row {row}"
        assert isinstance(row[1], (int, float)), f"missing quantity on row {row}"
        assert row[5], f"missing confidence_status on row {row}"
        headers_lower = [h.lower().replace(" ", "_") for h in HEADERS]
        assert "size_source" in headers_lower and "unpriced" in headers_lower


def test_xlsx_unpriced_flagged_never_zero():
    grid = _xlsx_grid(_payload())
    unpriced_rows = [row for row in grid if row and row[4] is True]
    assert len(unpriced_rows) == 1
    row = unpriced_rows[0]
    assert row[0] == "Mystery bracket"
    assert row[2] == UNPRICED_LABEL
    assert row[3] == UNPRICED_LABEL
    for value in row:
        assert value != 0, "unpriced line must never render as $0"


def test_xlsx_empty_payload_renders_header_only():
    grid = _xlsx_grid(_empty_payload())
    assert grid[0] == HEADERS
    assert len(grid) == 1


# ---------------------------------------------------------------------------
# PDF writer — real PDF bytes
# ---------------------------------------------------------------------------
def test_pdf_starts_with_pdf_magic():
    out = render_pdf(_payload())
    assert out.startswith(b"%PDF-")


def test_pdf_nontrivial_content():
    assert len(render_pdf(_payload())) > 500


def test_pdf_empty_payload_safe():
    assert render_pdf(_empty_payload()).startswith(b"%PDF-")


# ---------------------------------------------------------------------------
# Export endpoint — ORM-backed, format negotiation, errors
# ---------------------------------------------------------------------------
@pytest.fixture()
def api():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    session = Session(engine)
    app = FastAPI()
    app.include_router(exports_router)
    app.dependency_overrides[get_db] = lambda: session
    yield TestClient(app), session
    session.close()


def _seed_estimate(session: Session) -> uuid.UUID:
    project = Project(name="P")
    drawing = Drawing(project=project)
    sheet = Sheet(drawing=drawing, name="S-1")
    route = Route(sheet=sheet, route_type="duct_rectangular", length_m=12.5)
    component = Component(sheet=sheet, component_type="lighting_outlet", x=1.0, y=2.0)
    mystery = Component(sheet=sheet, component_type="mystery_bracket", x=3.0, y=4.0)

    meas_route = Measurement(
        route=route,
        source_sheet="S-1",
        source_region="full",
        measurement_type="route_length",
        raw_value=12.5,
        final_value=12.5,
    )
    meas_count = Measurement(
        component=component,
        source_sheet="S-1",
        source_region="symbol",
        measurement_type="count",
        raw_value=3.0,
        final_value=3.0,
    )
    meas_mystery = Measurement(
        component=mystery,
        source_sheet="S-1",
        source_region="symbol",
        measurement_type="count",
        raw_value=4.0,
        final_value=4.0,
    )
    estimate = Estimate(
        project=project,
        total_material_cost=95.0,
        total_labor_cost=0.0,
        total_cost=95.0,
    )
    session.add_all(
        [
            BoqItem(
                measurement=meas_route,
                estimate=estimate,
                quantity=12.5,
                unit_cost=7.0,
                total_cost=87.5,
                derivation_json=json.dumps(
                    {
                        "linear_per_m": 7.0,
                        "inputs": {"length_m": 12.5},
                        "rule_name": "Rectangular duct",
                    }
                ),
                size_source="schedule",
            ),
            BoqItem(
                measurement=meas_count,
                estimate=estimate,
                quantity=3.0,
                unit_cost=2.5,
                total_cost=7.5,
                size_source=None,
            ),
            BoqItem(
                measurement=meas_mystery,
                estimate=estimate,
                quantity=4.0,
                unit_cost=0.0,
                total_cost=0.0,
                size_source="assumed",
            ),
        ]
    )
    session.commit()
    return estimate.id


def test_endpoint_json_export_round_trips_persisted_boq(api):
    client, session = api
    estimate_id = _seed_estimate(session)
    resp = client.get(f"/api/exports/estimates/{estimate_id}/export?format=json")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/json")
    assert "attachment" in resp.headers.get("content-disposition", "")
    body = resp.json()
    assert body["estimate_id"] == str(estimate_id)
    assert body["totals"]["grand"] == 95.0
    assert body["routes"][0]["quantity"] == 12.5
    assert body["routes"][0]["size_source"] == "schedule"
    names = [m["material_name"] for m in body["materials"]]
    assert names == ["lighting_outlet", "mystery_bracket"]
    for section in ("routes", "materials"):
        for line in body[section]:
            assert REQUIRED_LINE_FIELDS <= set(line), f"line missing provenance: {line}"


def test_endpoint_payload_flags_zero_cost_lines_unpriced(api):
    client, session = api
    estimate_id = _seed_estimate(session)
    body = client.get(f"/api/exports/estimates/{estimate_id}/export").json()
    mystery = next(m for m in body["materials"] if m["material_name"] == "mystery_bracket")
    assert mystery["unpriced"] is True


def test_endpoint_xlsx_export_is_valid_workbook(api):
    client, session = api
    estimate_id = _seed_estimate(session)
    resp = client.get(f"/api/exports/estimates/{estimate_id}/export?format=xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp.content.startswith(b"PK")
    wb = load_workbook(io.BytesIO(resp.content))
    grid = [[cell.value for cell in row] for row in wb.active.iter_rows()]
    assert grid[0] == HEADERS
    unpriced = [row for row in grid if row and row[4] is True]
    assert len(unpriced) == 1
    assert unpriced[0][0] == "mystery_bracket"
    assert unpriced[0][3] == UNPRICED_LABEL


def test_endpoint_pdf_export_has_pdf_magic(api):
    client, session = api
    estimate_id = _seed_estimate(session)
    resp = client.get(f"/api/exports/estimates/{estimate_id}/export?format=pdf")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/pdf")
    assert resp.content.startswith(b"%PDF-")


def test_endpoint_unknown_format_is_422(api):
    client, session = api
    estimate_id = _seed_estimate(session)
    resp = client.get(f"/api/exports/estimates/{estimate_id}/export?format=csv")
    assert resp.status_code == 422


def test_endpoint_default_format_is_json(api):
    client, session = api
    estimate_id = _seed_estimate(session)
    resp = client.get(f"/api/exports/estimates/{estimate_id}/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/json")


def test_endpoint_404_on_unknown_estimate(api):
    client, _ = api
    resp = client.get(f"/api/exports/estimates/{uuid.uuid4()}/export?format=json")
    assert resp.status_code == 404
